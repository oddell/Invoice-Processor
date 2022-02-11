
import os
import sys
import pandas as pd
import shutil
import csv

from PyPDF2 import PdfFileWriter, PdfFileReader

from google.cloud import documentai

from openpyxl import load_workbook

os.chdir(r"M:\Contracts Folder")
applicationPath = r"M:\Contracts Folder\Utilities\Invoice Processor"

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = applicationPath+'\\bright-velocity-333609-6081a4699c11.json'

project_id = "bright-velocity-333609"
location = "eu"  # Format is 'us' or 'eu'
processor_id = "fab48bf677935aeb"  # Create processor in Cloud Console
companyName = "mildren"

class Processor():
  
    """AI, information handling and organisation"""

    def __init__(self):
        #Class Preperation

        Processor.reject = False
        Processor.contractList = Utilities.FindContractList()

    def DoProcurementAI(project_id: str, location: str, processor_id: str, file_path: str):

        # Modified DocumentAI quickstart

        opts = {}
        if location == "eu":
            opts = {"api_endpoint": "eu-documentai.googleapis.com"}

        client = documentai.DocumentProcessorServiceClient(client_options=opts)

        # The full resource name of the processor, e.g.:
        # projects/project-id/locations/location/processor/processor-id

        name = f"projects/{project_id}/locations/{location}/processors/{processor_id}"

        # Read the file into memory
        with open(file_path, "rb") as image:
            image_content = image.read()

        document = {"content": image_content, "mime_type": "application/pdf"}

        # Configure the process request
        request = {"name": name, "raw_document": document}

        result = client.process_document(request=request)
        
        document = result.document
        entities = document.entities

        # Prepare attributes
        types = []
        values = []
        confidence = []
        normalizedValues = []

        # Read the text recognition output from the processor
        for entity in entities:
            types.append(entity.type_)
            values.append(entity.mention_text)
            confidence.append(round(entity.confidence,4))
            normalizedValues.append(entity.normalized_value.text)

        # Prepare dictionary of info -> MAKE INTO OBJECT  
        invoiceInfo = dict(zip(types, values))
        invoiceNormalizedInfo = dict(zip(types, normalizedValues))

        #Populate with normalized values if possible
        for key in invoiceNormalizedInfo:
            if invoiceNormalizedInfo[key] != "":
                invoiceInfo[key] = invoiceNormalizedInfo[key]

        return(invoiceInfo)

    def FindContractNumber(invoiceInfo):

        #Find current projects -> MAKE CURRENT -ACTIVE- PROJECTS
        projectList = Utilities.contractList
        found = False
        for project in projectList:
            for information in invoiceInfo.values():
                if project in information:
                    invoiceInfo["project_no"] = project
                    found = True
                    break
            else:
                continue
            break
        if found == False:
            print("Project Number Not Found")
        else:
            print(invoiceInfo["project_no"])
        return invoiceInfo

    def TroubleshootInfo(invoiceInfo):
        #Prepare wanted info:
        keys = ("invoice_date", "supplier_name", "invoice_id", "line_item", "net_amount","project_no")
        invoiceInfo["IsHire"] = "Purchase"
        chars = "£Ł\\,L"
        for key in keys:
            try:
                invoiceInfo[key]
            except KeyError:
                found = False

                #Find Net Amount from Tax - Vat
                if key == "net_amount":
                    try:
                        #Remove characters that look like £
                        for c in chars:
                            invoiceInfo['total_amount'] = invoiceInfo['total_amount'].replace(c, "")
                            invoiceInfo['total_tax_amount'] = invoiceInfo['total_tax_amount'].replace(c,"")
                        invoiceInfo['net_amount'] = float(invoiceInfo['total_amount']) - float(invoiceInfo['total_tax_amount'])
                        try:
                            print("Calculated Net of "+str(invoiceInfo['total_amount'])+" - "+str(invoiceInfo['total_tax_amount'])+" = "+str(invoiceInfo['net_amount'])+" for invoice "+invoiceInfo["invoice_id"])
                            found = True
                        except KeyError:
                            found = False
                            
                    except:
                        continue
                
                #If no ID use Order Number
                if  key == "invoice_id":
                    try:
                        invoiceInfo[key] = invoiceInfo["purchase_order"]
                        print("Invoice ID Altered to "+ invoiceInfo[key])
                        found = True
                    except KeyError:
                        continue
                if found == False:
                    print("Error: "+key+" not found!")
                    invoiceInfo[key] = "Error"
                    Processor.reject = True

                
        #Supplier name can not be company name
        if companyName in str(invoiceInfo["supplier_name"]).lower():
            invoiceInfo["supplier_name"] = "Error"
            Processor.reject = True
        
        #Tidy Net
        try:
            if invoiceInfo['net_amount']:   
                try:
                    for c in chars:
                        invoiceInfo['net_amount'] = invoiceInfo['net_amount'].replace(c, "")
                except:
                    ""
        except KeyError:
            Processor.reject = True

        #Standardise Date
        try:
            invoiceInfo['excel_date'] = (pd.to_datetime(invoiceInfo['invoice_date'], dayfirst=True))
            invoiceInfo['invoice_date'] = (pd.to_datetime(invoiceInfo['excel_date'], dayfirst=True)).strftime("%b %Y")
            
        except:
            invoiceInfo['excel_date'] = "Error"
            Processor.reject = True

        #Strip new lines
        try:
            invoiceInfo["invoice_id"]=invoiceInfo["invoice_id"].replace("\n","")
        except:
            pass

        #Check if hire invoice
        for information in invoiceInfo.values():
            if "hire" in str(information).lower():
                invoiceInfo["IsHire"] = "Hire"
        return(invoiceInfo)
    
    def RejectPDF(invoiceInfo, inputPDFPath):
        
        #Prepare pdf and csv outputs
        rejectNumber = (len(os.listdir(applicationPath+"/Rejected PDFs/"))/2) + 1
        pdfOutputPath = applicationPath+"/Rejected PDFs/Reject "+str(rejectNumber)+".pdf"
        csvOutputPath = applicationPath+"/Rejected PDFs/Reject "+str(rejectNumber)+".csv"
        
        #Write InvoiceInfo to csv
        w = csv.writer(open(csvOutputPath, "w"))
        for key, val in invoiceInfo.items():
            try:
                w.writerow([key, val])
            except:
                pass
        
        print("Reached Reject #"+str(rejectNumber))

        #Move pdf to temp storage
        shutil.move(inputPDFPath, pdfOutputPath)



        pass

    def AcceptPDF(invoiceInfo, inputPDFPath):
        
        #Checks and Rectifies Contracts Local Excel
        filePath = Utilities.FindContractPath(invoiceInfo["project_no"])
        filePath += "/Commercial/CVR's/Invoice Consolidation"
        os.makedirs(filePath, exist_ok=True)
        localExcelPath = filePath+"/"+invoiceInfo["project_no"]+".xlsx"
        if os.path.isfile(localExcelPath) == False:
            print("Excel created at "+filePath)
            shutil.copy(applicationPath+"/GIR Template.xlsx", localExcelPath)

        #Send invoice to archive
        pdfOutputPath = filePath+"/Invoice Archive/"+invoiceInfo['invoice_date']+"/"+invoiceInfo['invoice_id']+".pdf"
        os.makedirs(os.path.dirname(pdfOutputPath), exist_ok=True)
        shutil.move(inputPDFPath, pdfOutputPath)

        #Save to Contract Local Excel
        wb = load_workbook(localExcelPath)
        ws = wb.active
        ws.append([invoiceInfo["excel_date"], invoiceInfo["supplier_name"] , '=HYPERLINK("{}", "{}")'.format(pdfOutputPath, invoiceInfo["invoice_id"]), invoiceInfo["IsHire"], invoiceInfo["line_item"] ,float(invoiceInfo["net_amount"])])
        wb.save(localExcelPath)

        #Save to Contract Global Excel
        globalExcelPath = applicationPath+"/Global Invoice Reconcilliation.xlsx"
        wb = load_workbook(globalExcelPath)
        ws = wb.active
        ws.append([invoiceInfo["excel_date"], invoiceInfo["project_no"] , invoiceInfo["supplier_name"] , '=HYPERLINK("{}", "{}")'.format(pdfOutputPath, invoiceInfo["invoice_id"]), invoiceInfo["IsHire"], invoiceInfo["line_item"] ,float(invoiceInfo["net_amount"])])
        AttemptSave= True
        wb.save(globalExcelPath)

        print(invoiceInfo['invoice_id']+" Processed")
        pass

    def Process(invoiceInfo, inputPDFPath):
    
        #DoProcurementAI gives next
        Processor.reject = False

        invoiceInfo = Processor.FindContractNumber(invoiceInfo)
        invoiceInfo = Processor.TroubleshootInfo(invoiceInfo)

        if Processor.reject == True:
            Processor.RejectPDF(invoiceInfo, inputPDFPath)
        else:
            Processor.AcceptPDF(invoiceInfo, inputPDFPath)


class Utilities():

    """Generic business utilities"""

    def __init__(self) -> None:
        pass

    def FindContractPath(contractNumber):
        
        #Uses project number to find path in business drive
        filePath = "M:\Contracts Folder"
        
        for dirNames in os.listdir(filePath):             
            if contractNumber[0:2] in dirNames[0:2]:
                filePath += "/" + dirNames
        
        for dirNames in os.listdir(filePath):             
            if contractNumber[0:5] in dirNames[0:5]:
                filePath += "/" + dirNames
        return filePath
    
    def FindContractList():
        
        #Finds existing contracts -> Needs to be active contracts
        filePath = "M:\Contracts Folder"
        contractList = []
        for dirNames in os.listdir(filePath):             
            if dirNames[0:2].isnumeric() == True:
                for subDirNames in os.listdir(filePath+"/"+dirNames):
                    if subDirNames[0:5].isnumeric() == True:
                        try:
                            contractList.append(subDirNames[0:5])
                        except:
                            pass
        Utilities.contractList = contractList
    
    def SplitPDFs():
        #Splits large PDF into individual
        #Needs to allow user to reject reciepts before AI
        
        inputFolderPath = applicationPath+"/PDF Input/"
        inputPDFPathList = [inputFolderPath+f for f in os.listdir(inputFolderPath) if os.path.isfile(os.path.join(inputFolderPath,f))]

        for inputPDFPath in inputPDFPathList:
            with open(inputPDFPath, mode='rb') as r:
                inputpdf = PdfFileReader(r)
                print(inputPDFPath)
                for i in range(inputpdf.numPages):
                    output = PdfFileWriter()
                    output.addPage(inputpdf.getPage(i))
                    with open(inputFolderPath+"/"+inputPDFPath[len(inputFolderPath)+1:]+str(i)+".pdf", "wb") as outputStream:
                        output.write(outputStream)
            os.remove(inputPDFPath)

        return [inputFolderPath+f for f in os.listdir(inputFolderPath) if os.path.isfile(os.path.join(inputFolderPath,f))]
        


def Main():
    inputPDFPathList = Utilities.SplitPDFs()
    print(inputPDFPathList)
    Utilities.FindContractList()
    for inputPDFPath in inputPDFPathList:
        invoiceInfo = Processor.DoProcurementAI(project_id, location, processor_id, inputPDFPath)
        Processor.Process(invoiceInfo, inputPDFPath)
    

if __name__ == '__main__':
    sys.exit(Main())