from distutils import command
import os
from pickle import TRUE
import sys
import tkinter as tk
from pdf2image import convert_from_path
from PIL import ImageTk
import collections as ct 
import csv
from tkPDFViewer import tkPDFViewer as pdf
import shutil
from openpyxl import load_workbook


from App import showpdf

os.chdir(r"M:\Contracts Folder")
applicationPath = r"M:\Contracts Folder\Utilities\Invoice Processor"

class ManualProcessor():
    
    def __init__(self) -> None:
        pass

    def FindPDFs():

        rejectDictionary = {}
        rejectFolderPath = applicationPath+"/Rejected PDFs/"
        rejectPathList = [rejectFolderPath+f for f in os.listdir(rejectFolderPath) if os.path.isfile(os.path.join(rejectFolderPath,f))]
        for pdfFilePath in rejectPathList:
            if pdfFilePath.endswith(".pdf") == True:
                rejectDictionary[pdfFilePath.replace(".pdf",".csv")] = pdfFilePath
        print(rejectDictionary)

        return rejectDictionary
    
    def FindContractPath(contractNumber):
        
        filePath = "M:\Contracts Folder"
        
        for dirNames in os.listdir(filePath):             
            if contractNumber[0:2] in dirNames[0:2]:
                filePath += "/" + dirNames
        
        for dirNames in os.listdir(filePath):             
            if contractNumber[0:5] in dirNames[0:5]:
                filePath += "/" + dirNames
        return filePath

    def ShowPDF(PDF_PATH, key):
    
        try:
            ManualProcessor.previousEntry
        except:
            ManualProcessor.previousEntry = {}
        
        root = tk.Tk()
        
        root.title("Manual Entry")

        root.geometry('2500x1000')

        v1 = pdf.ShowPdf()
        v1.img_object_li.clear()

        v2 = v1.pdf_view(root, 
                 pdf_location = PDF_PATH,  
                 width = 80, height = 80) 
        
        v2.pack(side=tk.LEFT)

        typeLabel = tk.Label(root, text='Enter '+key)
        typeLabel.pack( side = tk.LEFT )

        ManualProcessor.entryBox = tk.Entry (root)
        ManualProcessor.entryBox.pack( side = tk.LEFT ) 

        def EnterButtonPress():
            ManualProcessor.entry = ManualProcessor.entryBox.get()
            ManualProcessor.previousEntry[key] = ManualProcessor.entry
            root.destroy()
            return 
            
        
        enterButton = tk.Button(root, text="Enter", command=EnterButtonPress)
        enterButton.pack( side = tk.LEFT )

        def RejectButtonPress():
            ManualProcessor.entry = "Reject"
            root.destroy()
            return
            

        rejectButton = tk.Button(root, text="Reject", command=RejectButtonPress)
        rejectButton.pack( side = tk.LEFT )

        def PreviousButtonPress():
            ManualProcessor.entry = ManualProcessor.previousEntry[key]
            root.destroy()
            return

        try:
            previousLabel = tk.Label(root, text= "Previous "+str(ManualProcessor.previousEntry[key]))
            previousLabel.pack( side = tk.LEFT ) 

            previousButton = tk.Button(root, text="Previous", command=PreviousButtonPress)
            previousButton.pack( side = tk.LEFT )
        except:
            pass

        
        

        root.mainloop()

    def AcceptPDF(invoiceInfo, inputPDFPath):
        
        #Checks and Rectifies Contracts Local Excel
        filePath = ManualProcessor.FindContractPath(invoiceInfo["project_no"])
        filePath += "/Commercial/CVR's/Invoice Consolidation"
        os.makedirs(filePath, exist_ok=True)
        localExcelPath = filePath+"/"+invoiceInfo["project_no"]+".xlsx"
        if os.path.isfile(localExcelPath) == False:
            print("Excel created at "+filePath)
            shutil.copy(applicationPath+"/GIR Template.xlsx", localExcelPath)

        pdfOutputPath = filePath+"/Invoice Archive/"+invoiceInfo['invoice_date']+"/"+invoiceInfo['invoice_id']+".pdf"


        #Save to Contract Local Excel
        wb = load_workbook(localExcelPath)
        ws = wb.active
        ws.append([invoiceInfo["excel_date"], invoiceInfo["supplier_name"] , '=HYPERLINK("{}", "{}")'.format(pdfOutputPath, invoiceInfo["invoice_id"]), invoiceInfo["IsHire"], invoiceInfo["line_item"] ,float(invoiceInfo["net_amount"])])
        wb.save(localExcelPath)

        #Save to Contract Global Excel
        globalExcelPath = applicationPath+"/Global Invoice Reconcilliation.xlsx"
        wb = load_workbook(globalExcelPath)
        ws = wb.active
        ws.append([invoiceInfo["excel_date"], invoiceInfo["project_no"] , invoiceInfo["supplier_name"] , '=HYPERLINK("{}", "{}")'.format(pdfOutputPath, invoiceInfo["invoice_id"]), invoiceInfo["IsHire"], invoiceInfo["line_item"] ,float(invoiceInfo["net_amount"])])
        AttemptSave= True
        wb.save(globalExcelPath)

        #Send invoice to archive
        os.makedirs(os.path.dirname(pdfOutputPath), exist_ok=True)
        shutil.move(inputPDFPath, pdfOutputPath)

        print(invoiceInfo['invoice_id']+" Processed")
        pass
    
    def RejectPDF():
        print("PDF REJECTED")

    def Process(pdfDictionary):
        invoiceInfo = {}

        for csvPath in pdfDictionary:
            reject = False
            invoiceInfo = {}
            
            with open(csvPath) as file:
                reader = csv.reader(file)
                for row in reader:
                    if reject == False:
                        try:
                            type, value = row
                            invoiceInfo[type] = value
                            if value == "Error":
                                ManualProcessor.ShowPDF(pdfDictionary[csvPath], type)
                                print(ManualProcessor.entry)
                                if ManualProcessor.entry == "Reject":
                                    reject = True      
                                else:
                                    invoiceInfo[type] = ManualProcessor.entry
                                
                        except:
                            pass
                        testKeys = ["excel_date","project_no","supplier_name","net_amount"]

                if reject == False:   
                    for key in testKeys:
                        try:
                            invoiceInfo[key]
                        except:
                            ManualProcessor.ShowPDF(pdfDictionary[csvPath], key)
                            if ManualProcessor.entry == "Reject":
                                reject = True      
                            else:
                                invoiceInfo[key] = ManualProcessor.entry
                        
                    

            if reject == False:
                
                ManualProcessor.AcceptPDF(invoiceInfo, pdfDictionary[csvPath])
                os.remove(csvPath)
            else:
                os.remove(csvPath)
                os.remove(pdfDictionary[csvPath])
                       

def Main():
    pdfDictionary = ManualProcessor.FindPDFs()
    ManualProcessor.Process(pdfDictionary)

if __name__ == '__main__':
    sys.exit(Main())